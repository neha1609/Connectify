import imaplib
import email
from email.header import decode_header
import webbrowser
import os

from openpyxl import Workbook
from openpyxl import load_workbook


# account credentials
username = "nehajhawar1609@gmail.com"
password = "gm@!l.c0m"
filename = "Email.txt"

def clean(text):
    # clean text for creating a folder
    return "".join(c if c.isalnum() else "_" for c in text)


filename = "Email_data.xlsx"
if os.path.isfile(filename)==False:
    wb = Workbook()
    sheet=wb.worksheets[0]
    sheet.cell(row=1,column=1).value="EMAIL"
    sheet.cell(row=1,column=2).value="PRODUCT"
    
    wb.save(filename = filename)
wb=load_workbook(filename)
sheet=wb.worksheets[0]
maxr=sheet.max_row

# create an IMAP4 class with SSL 
imap = imaplib.IMAP4_SSL("imap.gmail.com")
# authenticate
imap.login(username, password)

status, messages = imap.select("INBOX")
# number of top emails to fetch
N = 3
# total number of emails
messages = int(messages[0])
print(messages)

for i in range(messages, messages-N, -1):
    # fetch the email message by ID
    res, msg = imap.fetch(str(i), "(RFC822)")
    print(msg)
    for response in msg:

        if isinstance(response, tuple):
            # parse a bytes email into a message object
            msg = email.message_from_bytes(response[1])
            # decode the email subject
            subject, encoding = decode_header(msg["Subject"])[0]
            if isinstance(subject, bytes):
                # if it's a bytes, decode to str
                if(encoding!=None):
                    subject = subject.decode(encoding)
            # decode email sender
            From, encoding = decode_header(msg.get("From"))[0]
            if isinstance(From, bytes):
                From = From.decode(encoding)
            #print("Subject:", subject)
            print("From:", From)
            maxr=sheet.max_row
            print(maxr)
            sheet.cell(row=maxr+1,column=1).value=str(From)
            # if the email message is multipart
            if msg.is_multipart():
                # iterate over email parts
                for part in msg.walk():
                    # extract content type of email
                    content_type = part.get_content_type()
                    content_disposition = str(part.get("Content-Disposition"))
                    try:
                        # get the email body
                        body = part.get_payload(decode=True).decode()

                    except:
                        pass
                    if content_type == "text/plain" and "attachment" not in content_disposition:
                        # print text/plain emails and skip attachments
                        #print(body)
                        '''folder_name = clean(subject)
                        print(folder_name)
                        file = 'myfile.txt'
                        fileName = 'C:/xampp/htdocs/EnggPros/' + folder_name+ '_myfile.txt'
                        file1 = open(fileName,"a") 
                    
                        #file1 = open("myfile.txt","a") 
                        file1.write("Hello \n") 
                        
                        #maxr=sheet.max_row'''
                        try:
                            #file1.writelines(body) 
                            body=body.splitlines()
                            sheet.cell(row=maxr+1,column=2).value=body[0]
                            wb.save(filename = filename)
                        except:
                            pass
                        #file1.close()
                    elif "attachment" in content_disposition:
                        # download attachment
                        filename = part.get_filename()
                        if filename:
                            folder_name = clean(subject)
                            if not os.path.isdir(folder_name):
                                # make a folder for this email (named after the subject)
                                os.mkdir(folder_name)
                            filepath = os.path.join(folder_name, filename)
                            # download attachment and save it
                            open(filepath, "wb").write(part.get_payload(decode=True))
            else:
                # extract content type of email
                content_type = msg.get_content_type()
                # get the email body
                body = msg.get_payload(decode=True).decode()
                if content_type == "text/plain":
                    # print only text email parts
                    #print(body)
                    '''print(folder_name)
                    fileName = 'C:/xampp/htdocs/EnggPros/' + folder_name + '_myfile.txt'
                    file1 = open(fileName,"a") 
                    file1.write("Hello \n") 
                    file1.writelines(body) 
                    file1.close()'''
                    body=body.splitlines()
                    #maxr=sheet.max_row
                    sheet.cell(row=maxr+1,column=2).value=body[0]
                    wb.save(filename = filename)
                    
            if content_type == "text/html":
                # if it's HTML, create a new HTML file and open it in browser
                folder_name = clean(subject)
                if not os.path.isdir(folder_name):
                    # make a folder for this email (named after the subject)
                    os.mkdir(folder_name)
                filename = "index.html"
                filepath = os.path.join(folder_name, filename)
                # write the file
                open(filepath, "w", encoding='utf-8').write(body)
                # open in the default browser
                webbrowser.open(filepath)
            print("="*100)
        #wb.save(filename = filename)
# close the connection and logout
imap.close()
imap.logout()